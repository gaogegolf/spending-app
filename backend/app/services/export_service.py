"""Service for exporting reports in various formats (CSV, Excel, PDF)."""

import io
import csv
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, Any, List, Optional

from sqlalchemy.orm import Session

from app.models.transaction import Transaction
from app.models.account import Account
from app.services.stats_service import StatsService


class ExportService:
    """Service for generating export reports."""

    def __init__(self, db: Session):
        """Initialize export service.

        Args:
            db: Database session
        """
        self.db = db
        self.stats_service = StatsService(db)

    def export_to_csv(
        self,
        user_id: str = "default_user",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        account_id: Optional[str] = None,
        include_summary: bool = True
    ) -> io.StringIO:
        """Export report to CSV format.

        Args:
            user_id: User ID
            start_date: Start date filter
            end_date: End date filter
            account_id: Account filter
            include_summary: Whether to include summary section

        Returns:
            StringIO buffer with CSV content
        """
        output = io.StringIO()
        writer = csv.writer(output)

        # Default date range to current year if not specified
        if not start_date:
            start_date = date(datetime.now().year, 1, 1)
        if not end_date:
            end_date = date.today()

        # Write header
        writer.writerow(['Personal Finance Report'])
        writer.writerow([f'Period: {start_date} to {end_date}'])
        writer.writerow([f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}'])
        writer.writerow([])

        # Summary section
        if include_summary:
            summary = self.stats_service.get_date_range_summary(
                start_date, end_date, user_id, account_id
            )

            writer.writerow(['SUMMARY'])
            writer.writerow(['Total Spending', f'${summary["total_spend"]:.2f}'])
            writer.writerow(['Total Income', f'${summary["total_income"]:.2f}'])
            writer.writerow(['Net', f'${summary["net"]:.2f}'])
            writer.writerow(['Transaction Count', summary["transaction_count"]])
            writer.writerow([])

            # Category breakdown
            writer.writerow(['SPENDING BY CATEGORY'])
            writer.writerow(['Category', 'Amount', 'Count', 'Percentage'])
            for cat in summary['category_breakdown']:
                writer.writerow([
                    cat['category'],
                    f'${cat["amount"]:.2f}',
                    cat['count'],
                    f'{cat["percentage"]:.1f}%'
                ])
            writer.writerow([])

        # Transaction details
        writer.writerow(['TRANSACTIONS'])
        writer.writerow([
            'Date', 'Merchant', 'Description', 'Amount', 'Type',
            'Category', 'Account', 'Is Spending'
        ])

        # Query transactions
        query = self.db.query(Transaction).join(Account).filter(
            Account.user_id == user_id,
            Transaction.date >= start_date,
            Transaction.date <= end_date
        )

        if account_id:
            query = query.filter(Transaction.account_id == account_id)

        transactions = query.order_by(Transaction.date.desc()).all()

        for txn in transactions:
            writer.writerow([
                txn.date.strftime('%Y-%m-%d'),
                txn.merchant_normalized or '',
                txn.description_raw[:50],
                f'${float(txn.amount):.2f}',
                txn.transaction_type.value,
                txn.category or '',
                txn.account.name if txn.account else '',
                'Yes' if txn.is_spend else 'No'
            ])

        output.seek(0)
        return output

    def export_to_excel(
        self,
        user_id: str = "default_user",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        account_id: Optional[str] = None
    ) -> io.BytesIO:
        """Export report to Excel format with multiple sheets.

        Args:
            user_id: User ID
            start_date: Start date filter
            end_date: End date filter
            account_id: Account filter

        Returns:
            BytesIO buffer with Excel content
        """
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows

        # Default date range
        if not start_date:
            start_date = date(datetime.now().year, 1, 1)
        if not end_date:
            end_date = date.today()

        output = io.BytesIO()
        wb = Workbook()

        # Styles
        header_font = Font(bold=True, size=12)
        title_font = Font(bold=True, size=14)
        money_format = '#,##0.00'

        # === Summary Sheet ===
        ws_summary = wb.active
        ws_summary.title = "Summary"

        summary = self.stats_service.get_date_range_summary(
            start_date, end_date, user_id, account_id
        )

        ws_summary['A1'] = "Personal Finance Report"
        ws_summary['A1'].font = title_font
        ws_summary['A2'] = f"Period: {start_date} to {end_date}"
        ws_summary['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        ws_summary['A5'] = "Summary"
        ws_summary['A5'].font = header_font
        ws_summary['A6'] = "Total Spending"
        ws_summary['B6'] = summary['total_spend']
        ws_summary['B6'].number_format = money_format
        ws_summary['A7'] = "Total Income"
        ws_summary['B7'] = summary['total_income']
        ws_summary['B7'].number_format = money_format
        ws_summary['A8'] = "Net"
        ws_summary['B8'] = summary['net']
        ws_summary['B8'].number_format = money_format
        ws_summary['A9'] = "Transaction Count"
        ws_summary['B9'] = summary['transaction_count']

        # === Category Breakdown Sheet ===
        ws_categories = wb.create_sheet("Categories")

        ws_categories['A1'] = "Spending by Category"
        ws_categories['A1'].font = title_font

        headers = ['Category', 'Amount', 'Count', 'Percentage']
        for col, header in enumerate(headers, 1):
            cell = ws_categories.cell(row=3, column=col, value=header)
            cell.font = header_font

        for row, cat in enumerate(summary['category_breakdown'], 4):
            ws_categories.cell(row=row, column=1, value=cat['category'])
            ws_categories.cell(row=row, column=2, value=cat['amount']).number_format = money_format
            ws_categories.cell(row=row, column=3, value=cat['count'])
            ws_categories.cell(row=row, column=4, value=cat['percentage'] / 100).number_format = '0.0%'

        # === Transactions Sheet ===
        ws_transactions = wb.create_sheet("Transactions")

        query = self.db.query(Transaction).join(Account).filter(
            Account.user_id == user_id,
            Transaction.date >= start_date,
            Transaction.date <= end_date
        )

        if account_id:
            query = query.filter(Transaction.account_id == account_id)

        transactions = query.order_by(Transaction.date.desc()).all()

        # Headers
        txn_headers = ['Date', 'Merchant', 'Description', 'Amount', 'Type', 'Category', 'Account', 'Is Spending']
        for col, header in enumerate(txn_headers, 1):
            cell = ws_transactions.cell(row=1, column=col, value=header)
            cell.font = header_font

        # Data
        for row, txn in enumerate(transactions, 2):
            ws_transactions.cell(row=row, column=1, value=txn.date)
            ws_transactions.cell(row=row, column=2, value=txn.merchant_normalized or '')
            ws_transactions.cell(row=row, column=3, value=txn.description_raw[:50])
            ws_transactions.cell(row=row, column=4, value=float(txn.amount)).number_format = money_format
            ws_transactions.cell(row=row, column=5, value=txn.transaction_type.value)
            ws_transactions.cell(row=row, column=6, value=txn.category or '')
            ws_transactions.cell(row=row, column=7, value=txn.account.name if txn.account else '')
            ws_transactions.cell(row=row, column=8, value='Yes' if txn.is_spend else 'No')

        # === Monthly Trends Sheet ===
        ws_monthly = wb.create_sheet("Monthly Trends")

        ws_monthly['A1'] = "Monthly Spending Trends"
        ws_monthly['A1'].font = title_font

        year = start_date.year
        yearly = self.stats_service.get_yearly_summary(year, user_id, account_id)

        monthly_headers = ['Month', 'Spending', 'Income', 'Net']
        for col, header in enumerate(monthly_headers, 1):
            cell = ws_monthly.cell(row=3, column=col, value=header)
            cell.font = header_font

        for row, m in enumerate(yearly['monthly_data'], 4):
            ws_monthly.cell(row=row, column=1, value=m['month_name'])
            ws_monthly.cell(row=row, column=2, value=m['total_spend']).number_format = money_format
            ws_monthly.cell(row=row, column=3, value=m['total_income']).number_format = money_format
            ws_monthly.cell(row=row, column=4, value=m['net']).number_format = money_format

        # Auto-adjust column widths
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(output)
        output.seek(0)
        return output

    def export_to_pdf(
        self,
        user_id: str = "default_user",
        start_date: Optional[date] = None,
        end_date: Optional[date] = None,
        account_id: Optional[str] = None
    ) -> io.BytesIO:
        """Export report to PDF format.

        Args:
            user_id: User ID
            start_date: Start date filter
            end_date: End date filter
            account_id: Account filter

        Returns:
            BytesIO buffer with PDF content
        """
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.graphics.shapes import Drawing
        from reportlab.graphics.charts.barcharts import VerticalBarChart

        # Default date range
        if not start_date:
            start_date = date(datetime.now().year, 1, 1)
        if not end_date:
            end_date = date.today()

        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=letter, topMargin=0.5*inch, bottomMargin=0.5*inch)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'Heading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=20
        )

        elements = []

        # Title
        elements.append(Paragraph("Personal Finance Report", title_style))
        elements.append(Paragraph(f"Period: {start_date} to {end_date}", styles['Normal']))
        elements.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']))
        elements.append(Spacer(1, 20))

        # Summary
        summary = self.stats_service.get_date_range_summary(
            start_date, end_date, user_id, account_id
        )

        elements.append(Paragraph("Summary", heading_style))

        summary_data = [
            ['Metric', 'Value'],
            ['Total Spending', f'${summary["total_spend"]:,.2f}'],
            ['Total Income', f'${summary["total_income"]:,.2f}'],
            ['Net', f'${summary["net"]:,.2f}'],
            ['Transaction Count', str(summary["transaction_count"])],
        ]

        summary_table = Table(summary_data, colWidths=[2*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F3F4F6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.white),
        ]))
        elements.append(summary_table)

        # Category Breakdown
        if summary['category_breakdown']:
            elements.append(Paragraph("Spending by Category", heading_style))

            cat_data = [['Category', 'Amount', 'Count', '%']]
            for cat in summary['category_breakdown'][:15]:  # Top 15 categories
                cat_data.append([
                    cat['category'] or 'Uncategorized',
                    f'${cat["amount"]:,.2f}',
                    str(cat['count']),
                    f'{cat["percentage"]:.1f}%'
                ])

            cat_table = Table(cat_data, colWidths=[2.5*inch, 1.5*inch, 1*inch, 1*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F9FAFB')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ]))
            elements.append(cat_table)

        # Monthly Chart (simple text representation since charts can be complex)
        year = start_date.year
        yearly = self.stats_service.get_yearly_summary(year, user_id, account_id)

        elements.append(Paragraph(f"Monthly Spending - {year}", heading_style))

        monthly_data = [['Month', 'Spending', 'Income', 'Net']]
        for m in yearly['monthly_data']:
            monthly_data.append([
                m['month_name'][:3],
                f'${m["total_spend"]:,.2f}',
                f'${m["total_income"]:,.2f}',
                f'${m["net"]:,.2f}'
            ])

        monthly_table = Table(monthly_data, colWidths=[1*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        monthly_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F3F4F6')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ]))
        elements.append(monthly_table)

        doc.build(elements)
        output.seek(0)
        return output
